from zipfile import ZipFile
from openpyxl.reader.excel import load_workbook
from pypdf import PdfReader
from tests.conftest import archive


def test_check_pdf_file():
      with ZipFile(archive) as zf:
            reader = PdfReader(zf.open('file1.pdf'))
            page_count = len(reader.pages)
            print(page_count)

            page = reader.pages[0]
            text = page.extract_text()
            print(text)

            assert page_count == 1
            assert 'Банк России' in text


def test_check_xlsx_file():
    with ZipFile(archive) as zf:
        with zf.open('file2.xlsx') as file:
            workbook = load_workbook(file)
            sheet = workbook.active
            assert sheet.cell(row=2, column=4).value == 'Физ.лицо'
            assert sheet.max_row == 1794


def test_check_csv_file():
    with ZipFile(archive) as zf:
        with zf.open('file3.csv') as file:
            sheet = file.read().decode('utf-8-sig')
            assert 'FIX' in sheet